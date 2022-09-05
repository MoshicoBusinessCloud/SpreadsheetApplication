# ======================================================================================================================
# ----------------------------------------------------------------------------------------------------------------------
# ##### Project Automation with Python ====Project Introduction==== Project Automation with Python ######
# Learn how to work with Spreadsheets, read spreadsheet file and automate stuff. In this exercise, we are going to read
# info from a spreadsheet and list eash company with respective product count, for example: Exercise -- Resort:
#  {"AAA Company": 43, "BBB Company": 17, "CCC Company": 15}
# EX: List each company with prospect product count.
# EX: List product with inventory less than 10.
# EX: List each company with respective total inventory value.
# EX: Write to the spreadsheet: Calculate and write inventory value for each product into the spreadsheet.
# ----------------------------------------------------------------------------------------------------------------------
# ======================================================================================================================
# Exercise 1 === Exercise 1
# =============================================================
# Version[1]: This code count the number of rows on the spreadsheet
# ==============================================================

# import openpyxl
#
# inv_file = openpyxl.load_workbook("compound_inventory.xlsx")
# product_list = inv_file["Sheet1"]
#
# products_per_supplier = {}
#
# print(product_list.max_row)

# ======================================================================================================================
# ----------------------------------------------------------------------------------------------------------------------
# ##### Project Automation with Python ====Project Introduction==== Project Automation with Python ######
# ----------------------------------------------------------------------------------------------------------------------
# ======================================================================================================================
# Exercise 1 continue === Exercise 1 continue
# ============================================================================================================
# Version[2]: This version displayed the number companies on the spreadsheet and the number listing they have.
# In this version of the code, we are going to calculate the total number of inventory for each supplier
# =============================================================================================================
# import openpyxl
#
# inv_file = openpyxl.load_workbook("compound_inventory.xlsx")
# product_list = inv_file["Sheet1"]
#
# product_per_supplier = {}
#
# for product_row in range(3, product_list.max_row + 1):
#     supplier_name = product_list.cell(product_row, 4).value

# # Here we are going to calculate the total number of inventory for each supplier
#     if supplier_name in product_per_supplier:
#         # current_num_products = product_per_supplier[supplier_name] the line bellow can be writen this way as well.
#         current_num_products = product_per_supplier.get(supplier_name)
#         product_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         print("adding a new supplier")
#         product_per_supplier[supplier_name] = 1
#
# print(product_per_supplier)

# ======================================================================================================================
# ----------------------------------------------------------------------------------------------------------------------
# ##### Project Automation with Python ====Project Introduction==== Project Automation with Python ######
# ----------------------------------------------------------------------------------------------------------------------
# ======================================================================================================================
# Exercise 2 === Exercise 2
# ================================================================================================================
# Version[3]: In this version of the code, we are going to calculate the total value of inventory of each supplier
# =================================================================================================================
# import openpyxl
#
# inv_file = openpyxl.load_workbook("compound_inventory.xlsx")
# product_list = inv_file["Sheet1"]
#
# product_per_supplier = {}
# total_value_per_supplier = {}
#
# for product_row in range(3, product_list.max_row + 1):
#     supplier_name = product_list.cell(product_row, 4).value
#     inventory = product_list.cell(product_row, 2).value
#     price = product_list.cell(product_row, 3).value

# # Here we are going to calculate the total number of inventory for each supplier
#     if supplier_name in product_per_supplier:
#         current_num_products = product_per_supplier.get(supplier_name)
#         product_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         product_per_supplier[supplier_name] = 1
#
#
# # Here we are going to calculate the total value of inventory of each supplier
#     if supplier_name in total_value_per_supplier:
#         current_total_value = total_value_per_supplier.get(supplier_name)
#         total_value_per_supplier[supplier_name] = current_total_value + inventory * price
#     else:
#         total_value_per_supplier[supplier_name] = inventory * price
# print(product_per_supplier)
# print(total_value_per_supplier)

# ======================================================================================================================
# ----------------------------------------------------------------------------------------------------------------------
# ##### Project Automation with Python ====Project Introduction==== Project Automation with Python ######
# ----------------------------------------------------------------------------------------------------------------------
# ======================================================================================================================
# Exercise 3 === Exercise 3
# ================================================================================================================
# Version[4]: In this version of the code, we are going use logic to calculate product with inventory less than 500
# =================================================================================================================

# import openpyxl
#
# inv_file = openpyxl.load_workbook("compound_inventory.xlsx")
# product_list = inv_file["Sheet1"]
#
# product_per_supplier = {}
# total_value_per_supplier = {}
# product_under_500_inventory = {}
#
# for product_row in range(3, product_list.max_row + 1):
#     supplier_name = product_list.cell(product_row, 4).value
#     inventory = product_list.cell(product_row, 2).value
#     price = product_list.cell(product_row, 3).value
#     product_number = product_list.cell(product_row, 1).value
#
# # Here we are going to calculate the total number of inventory for each supplier
#     if supplier_name in product_per_supplier:
#         current_num_products = product_per_supplier.get(supplier_name)
#         product_per_supplier[supplier_name] = current_num_products + 1
#     else:
#         product_per_supplier[supplier_name] = 1
#
# # Here we are going to calculate the total value of inventory of each supplier
#     if supplier_name in total_value_per_supplier:
#         current_total_value = total_value_per_supplier.get(supplier_name)
#         total_value_per_supplier[supplier_name] = current_total_value + inventory * price
#     else:
#         total_value_per_supplier[supplier_name] = inventory * price
#
# # Here we are going use logic to calculate product with inventory less than 500
#     if inventory < 500:
#         product_under_500_inventory[product_number] = inventory
#
#
# print(product_under_500_inventory)
# print(product_per_supplier)
# print(total_value_per_supplier)

# ======================================================================================================================
# ----------------------------------------------------------------------------------------------------------------------
# ##### Project Automation with Python ====Project Introduction==== Project Automation with Python ######
# ----------------------------------------------------------------------------------------------------------------------
# ======================================================================================================================
# Exercise 4 === Exercise 4
# ===================================================================================================================
# Version[5]: In this version of the code, we are going to add the value for the total inventory price
# ===================================================================================================================

import openpyxl

inv_file = openpyxl.load_workbook("compound_inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_500_inventory = {}

for product_row in range(3, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

# Here we are going to calculate the total number of inventory for each supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

# Here we are going to calculate the total value of inventory of each supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# Here we are going use logic to calculate product with inventory less than 500
    if inventory < 500:
        product_under_500_inventory[product_number] = inventory


# Here we are going to add the value for the t0tal inventory price
    inventory_price.value = inventory * price
inv_file.save("inventory_with_total_value.xlsx")

print(product_under_500_inventory)
print(product_per_supplier)
print(total_value_per_supplier)
#======================================================================================================================
#======================================================================================================================